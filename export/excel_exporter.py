"""
export/excel_exporter.py — Professional Excel Report with LIVE FORMULAS
=======================================================================
Sheet 1: Summary — key results at a glance
Sheet 2: Seismic — NBC 105:2025 with live Excel formulas (inputs in blue)
Sheet 3: Beam — IS 456 beam design with live formulas
Sheet 4: Foundation — footing checks
Sheet 5: Notes & Disclaimer
"""
from __future__ import annotations
from datetime import datetime
import os
from constants import APP_NAME, APP_VERSION # type: ignore


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ── Colour palette ─────────────────────────────────────────────────────────────
C_TITLE_BG  = "1F5C99"
C_TITLE_FG  = "FFFFFF"
C_HEAD1_BG  = "2E75B6"
C_HEAD2_BG  = "DBEEF4"
C_OK_BG     = "E2EFDA"; C_OK_FG  = "375623"
C_FAIL_BG   = "FFDFD1"; C_FAIL_FG= "7F0000"
C_WARN_BG   = "FFEB9C"; C_WARN_FG= "9C5700"
C_INPUT_FG  = "0070C0"   # blue  — user-changeable inputs
C_FORMULA_FG= "375623"   # green — Excel formula cells (recalculate live)
C_BORDER    = "BFBFBF"
C_ALT_ROW   = "F5F9FF"


def _thin():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(ws, row, col, text, bg=C_HEAD1_BG, fg=C_TITLE_FG, bold=True, merge_to=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(name="Arial", size=10, bold=bold, color=fg)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = _thin()
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
    return c

def _inp(ws, row, col, value, fmt=None, align="right"):
    """Blue input cell — value the user can change."""
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", size=10, bold=True, color=C_INPUT_FG)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _thin()
    if fmt: c.number_format = fmt
    return c

def _fml(ws, row, col, formula, fmt=None, align="right"):
    """Green formula cell — recalculates live when inputs change."""
    c = ws.cell(row=row, column=col, value=formula)
    c.font      = Font(name="Arial", size=10, color=C_FORMULA_FG)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _thin()
    if fmt: c.number_format = fmt
    return c

def _lbl(ws, row, col, text, bold=False, bg=None, merge_to=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(name="Arial", size=10, bold=bold)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border    = _thin()
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
    return c

def _sec(ws, row, text, end_col=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(name="Arial", size=11, bold=True, color=C_TITLE_FG)
    c.fill      = PatternFill("solid", fgColor=C_TITLE_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border    = _thin()
    ws.row_dimensions[row].height = 18

def _ok(ws, row, col, ok):
    if ok is True:
        c = ws.cell(row=row, column=col, value="OK  ✓")
        c.font = Font(name="Arial", size=10, bold=True, color=C_OK_FG)
        c.fill = PatternFill("solid", fgColor=C_OK_BG)
    elif ok is False:
        c = ws.cell(row=row, column=col, value="REVISE  ✗")
        c.font = Font(name="Arial", size=10, bold=True, color=C_FAIL_FG)
        c.fill = PatternFill("solid", fgColor=C_FAIL_BG)
    else:
        c = ws.cell(row=row, column=col, value="—")
        c.font = Font(name="Arial", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = _thin()

def _proj_header(ws, sheet_title):
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"STRUCTURAL DESIGN CALCULATIONS  —  {sheet_title.upper()}"
    c.font  = Font(name="Arial", size=13, bold=True, color=C_TITLE_FG)
    c.fill  = PatternFill("solid", fgColor=C_TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 14
    ws.merge_cells("A3:H3")
    c3 = ws["A3"]
    c3.value = "Standards: NBC 105:2025 (Second Revision)  ·  IS 456:2000  ·  IS 875 Part 1 & 2"
    c3.font  = Font(name="Arial", size=9, italic=True, color="595959")
    c3.alignment = Alignment(horizontal="center")
    return 5   # first content row


def generate_excel_report(data: dict, output_path: str, mode: str = "detailed") -> str:
    if not OPENPYXL_OK:
        raise ImportError("openpyxl required. Run: pip install openpyxl")

    wb  = Workbook()
    pi  = data.get("project_info", {})
    seism = data.get("seismic", {})
    slab  = data.get("slab", {})
    beam  = data.get("beam", {})
    col   = data.get("column", {})
    fndg  = data.get("foundation", {})


    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — SUMMARY
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 16

    r = _proj_header(ws, "Summary Report")
    _sec(ws, r, "PROJECT INFORMATION", end_col=4); r+=1
    for lbl, val in [
        ("Project",         pi.get("project","—")),
        ("Engineer",        pi.get("engineer","—")),
        ("Checked By",      pi.get("checked_by","—")),
        ("Job No.",         pi.get("job_no","—")),
        ("Date",            pi.get("date", datetime.now().strftime("%Y-%m-%d"))),
        ("Standards",       "NBC 105:2025 · IS 456:2000 · IS 875"),
    ]:
        _lbl(ws, r, 1, lbl, bold=True); _lbl(ws, r, 2, val, merge_to=4); r+=1
    r+=1

    if seism:
        _sec(ws, r, "SEISMIC DESIGN — NBC 105:2025", end_col=4); r+=1
        _hdr(ws,r,1,"Parameter"); _hdr(ws,r,2,"Value"); _hdr(ws,r,3,"Clause"); _hdr(ws,r,4,"Status"); r+=1
        for i,(lbl,val,clause,ok) in enumerate([
            ("Zone Factor Z",         f"{seism.get('Z',0):.2f}",         "§4.1.4",     True),
            ("Importance Factor I",   f"{seism.get('I',0):.2f}",         "Table 4-4",  True),
            ("Design Period T [s]",   f"{seism.get('T',0):.4f}",         "§5.1.2–3",   True),
            ("Ch(T)",                 f"{seism.get('Ch_T',0):.4f}",      "§4.1.2",     True),
            ("C(T)",                  f"{seism.get('C_T',0):.4f}",       "§4.1.1",     True),
            ("Cs(T)",                 f"{seism.get('Cs_T',0):.4f}",      "§4.2",       True),
            ("Cv(T)",                 f"{seism.get('Cv_T',0):.4f}",      "§4.3",       True),
            ("Cd(T) ULS",             f"{seism.get('Cd_ULS',0):.4f}",    "§6.1.1",     True),
            ("Cd(T) SLS",             f"{seism.get('Cd_SLS',0):.4f}",    "§6.1.2",     True),
            ("kd (deflection scale)", f"{seism.get('kd',0):.2f}",        "Table 6-1",  True),
            ("Orientation",           "Parallel" if seism.get("is_parallel") else "Non-Parallel", "§3.6", None),
            ("Snow Load",             "Included" if seism.get("include_snow") else "Excluded", "§3.6", None),
        ]):
            bg = C_ALT_ROW if i%2==0 else None
            _lbl(ws,r,1,lbl,bold=True,bg=bg); _lbl(ws,r,2,val,bg=bg); _lbl(ws,r,3,clause,bg=bg); _ok(ws,r,4,ok)
            r+=1
        r+=1


    if slab:

        _sec(ws, r, "SLAB DESIGN — IS 456:2000 (TWO-WAY TAB)", end_col=4); r+=1
        _hdr(ws,r,1,"Parameter"); _hdr(ws,r,2,"Value"); _hdr(ws,r,3,"Note"); _hdr(ws,r,4,"Status"); r+=1
        ssum = slab.get("summary", {})
        slab_rows = [
            ("Ly/Lx Ratio", str(ssum.get("ratio", "—")), "Two-way if 1.0 to 2.0", None),
            ("Effective depth d [mm]", str(ssum.get("d_eff", "—")), "From current trial depth", None),
            ("Factored load wu [kN/m²]", str(ssum.get("wu", "—")), "ULS load", None),
            ("Ast,min [mm²]", str(ssum.get("astmin", "—")), "IS 456 minimum steel", None),
        ]
        for i,(lbl,val,note,ok) in enumerate(slab_rows):
            bg = C_ALT_ROW if i%2==0 else None
            _lbl(ws,r,1,lbl,bold=True,bg=bg); _lbl(ws,r,2,val,bg=bg); _lbl(ws,r,3,note,bg=bg); _ok(ws,r,4,ok)
            r += 1
        notes_txt = str(slab.get("notes", "")).strip()
        if notes_txt:
            _lbl(ws, r, 1, "Design Notes", bold=True)
            _lbl(ws, r, 2, notes_txt, merge_to=4)
            r += 1
        r += 1

    if beam:
        _sec(ws, r, "BEAM DESIGN — NBC 105:2025 PRIORITY + IS 456 FALLBACK", end_col=4); r+=1

        _hdr(ws,r,1,"Check"); _hdr(ws,r,2,"Value"); _hdr(ws,r,3,"Clause"); _hdr(ws,r,4,"Status"); r+=1
        for i,(lbl,val,clause,ok) in enumerate([
            ("Code basis",            str(beam.get("code_design_basis", "NBC 105:2025 priority + IS fallback")), "—", None),
            ("Section b×D",           f"{beam.get('b','?')}×{beam.get('D','?')} mm",    "—",        None),
            ("Main/Comp dia",         f"Ø{beam.get('main_dia','?')}/Ø{beam.get('comp_dia','?')} mm", "NBC Annex A §4.1.2", True),
            ("Load wD/wL",            f"{beam.get('dl_kNm',0):.2f}/{beam.get('ll_kNm',0):.2f} kN/m", "Input", None),
            ("Mu,lim",                f"{beam.get('Mu_lim_kNm',0):.2f} kN·m",           "IS fallback §38.1",    True),
            ("Design Mu",             f"{beam.get('Mu_design_kNm',0):.2f} kN·m",        "IS fallback §22.2",    True),
            ("Section type",          "Doubly" if beam.get("is_doubly") else "Singly",  "IS Annex G",  True),
            ("Ast required",          f"{beam.get('Ast_req_mm2',0):.0f} mm²",           "IS §26.5.1",  True),
            ("Ast provided",          f"{beam.get('Ast_prov_mm2',0):.0f} mm²",          "IS §26.5.1",
             beam.get("Ast_prov_mm2",0) >= beam.get("Ast_req_mm2",0)),
            ("Shear status",          beam.get("shear",{}).get("status","—"),            "IS §40 + NBC Annex A",      None),
            ("Stirrups @ support",    f"{(beam.get('shear',{}).get('Sv_end_zone_user_mm') or beam.get('shear',{}).get('Sv_end_zone_mm') or beam.get('shear',{}).get('Sv_mm') or 0):.0f} mm c/c", "NBC Annex A §4.1.3", True),
            ("Stirrups @ mid/main",   f"{(beam.get('shear',{}).get('Sv_mid_zone_user_mm') or beam.get('shear',{}).get('Sv_mid_zone_mm') or beam.get('shear',{}).get('Sv_mm') or 0):.0f} mm c/c", "NBC Annex A §4.1.3", True),
            ("Dev. length Ld",        f"{beam.get('Ld_mm',0):.0f} mm",                  "IS fallback §26.2.1",  True),

        ]):

            bg = C_ALT_ROW if i%2==0 else None
            _lbl(ws,r,1,lbl,bold=True,bg=bg); _lbl(ws,r,2,val,bg=bg); _lbl(ws,r,3,clause,bg=bg); _ok(ws,r,4,ok)
            r+=1
        bnotes = beam.get("notes", [])
        if bnotes:
            _lbl(ws, r, 1, "Beam design notes", bold=True)
            _lbl(ws, r, 2, " | ".join(str(n) for n in bnotes[:6]), merge_to=4)
            r += 1
        r+=1


    if col:
        _sec(ws, r, "COLUMN DESIGN — IS 456:2000 + NBC 105 Annex A", end_col=4); r+=1
        _hdr(ws,r,1,"Check"); _hdr(ws,r,2,"Value"); _hdr(ws,r,3,"Clause"); _hdr(ws,r,4,"Status"); r+=1
        for i,(lbl,val,clause,ok) in enumerate([
            ("Section b×D",             f"{col.get('b_mm','?')}×{col.get('D_mm','?')} mm",  "—",      None),
            ("Slenderness λx / λy",      f"{col.get('lambda_x',0):.2f} / {col.get('lambda_y',0):.2f}", "§25.1.2", not col.get('is_slender')),
            ("Pure Axial Capacity Pu,max", f"{col.get('Pu_max_kN',0):.1f} kN", "§39.3", col.get('Pu_max_kN',0)>=col.get('Pu_kN',0)),
            ("Biaxial interaction",      f"{col.get('interaction',0):.4f}",                   "§39.6",  col.get("interaction",1)<=1.0),
            ("Steel %",                  f"{col.get('steel_pct',0):.2f}%",                    "§26.5.3",0.8<=col.get("steel_pct",0)<=4.0),
            ("Bars provided",            f"{col.get('no_of_bars',0)}×Ø{int(col.get('bar_dia_mm',0))}mm","—",None),
            ("Tie spacing",              f"{col.get('tie_spacing_mm',0):.0f} mm c/c",          "§26.5.3",True),
            ("NBC 105 Confinement zone", f"{col.get('conf_zone_mm',0):.0f} mm",                "Annex A",True),
            ("Confinement tie spacing",  f"{col.get('conf_tie_sp_mm',0):.0f} mm c/c",          "Annex A",True),
            ("Ash check",               "OK ✓" if col.get("hoop_ok") else "FAIL ✗",             "Annex A",col.get("hoop_ok")),
        ]):
            bg = C_ALT_ROW if i%2==0 else None
            _lbl(ws,r,1,lbl,bold=True,bg=bg); _lbl(ws,r,2,val,bg=bg); _lbl(ws,r,3,clause,bg=bg); _ok(ws,r,4,ok)
            r+=1
        r+=1


    if fndg:

        _sec(ws, r, "FOUNDATION DESIGN — IS 456:2000 §34", end_col=4); r+=1
        _hdr(ws,r,1,"Check"); _hdr(ws,r,2,"Value"); _hdr(ws,r,3,"Clause"); _hdr(ws,r,4,"Status"); r+=1
        for i,(lbl,val,clause,ok) in enumerate([
            ("Plan size",          f"{fndg.get('L_mm',0)}×{fndg.get('B_mm',0)} mm",              "§34.1",   None),
            ("Depth D / d",        f"{fndg.get('D_mm',0):.0f} / {fndg.get('d_mm',0):.0f} mm",    "§34.1.3", fndg.get("D_mm",0)>=300),
            ("q_max (service)",    f"{fndg.get('q_max_kPa',0):.2f} kN/m²",                        "§34.2",   fndg.get("pressure_ok")),
            ("q_min (service)",    f"{fndg.get('q_min_kPa',0):.2f} kN/m²",                        "§34.2.4", fndg.get("q_min_kPa",0)>=0),
            ("One-way shear",      f"τv={fndg.get('tau_v_L',0):.3f} MPa",                         "§34.4.2a",fndg.get("one_way_ok")),
            ("Punching shear",     f"τv={fndg.get('tau_v_punch',0):.3f} MPa",                     "§34.4.2b",fndg.get("punch_ok")),
            ("Development length", f"Ld={fndg.get('Ld_mm',0):.0f} mm",                            "§34.4.3", fndg.get("dev_ok")),
        ]):
            bg = C_ALT_ROW if i%2==0 else None
            _lbl(ws,r,1,lbl,bold=True,bg=bg); _lbl(ws,r,2,val,bg=bg); _lbl(ws,r,3,clause,bg=bg); _ok(ws,r,4,ok)
            r+=1

    ws.freeze_panes = "A6"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — SEISMIC with LIVE EXCEL FORMULAS
    # ══════════════════════════════════════════════════════════════════════════
    if seism and mode == "detailed":
        ws2 = wb.create_sheet("Seismic (NBC 105)")
        for col_l, w in [("A",42),("B",22),("C",18),("D",18),("E",35)]:
            ws2.column_dimensions[col_l].width = w

        r = _proj_header(ws2, "Seismic Design — NBC 105:2025 (Live Formulas)")
        # Legend
        ws2.merge_cells(f"A{r}:E{r}")
        lc = ws2.cell(row=r, column=1,
            value="BLUE cells = inputs (edit to recalculate)   |   GREEN cells = Excel formulas (auto-recalculate)")
        lc.font = Font(name="Arial", size=9, italic=True, color="595959"); r+=1

        # ── SECTION A: Site Parameters ─────────────────────────────────────
        _sec(ws2, r, "A. SITE & STRUCTURAL PARAMETERS  (edit blue cells)", end_col=5); r+=1
        _hdr(ws2,r,1,"Parameter"); _hdr(ws2,r,2,"Input Value"); _hdr(ws2,r,3,"Unit")
        _hdr(ws2,r,4,"Note"); _hdr(ws2,r,5,"Clause"); r+=1

        # Store row refs for formula use
        ROW = {}
        for key, label, val, unit, note, clause in [
            ("Z",  "Zone Factor Z",          seism.get("Z",0.36),   "—",  "Peak ground acc.", "§4.1.4"),
            ("I",  "Importance Factor I",    seism.get("I",1.0),    "—",  "Table 4-4",        "§4.1.5"),
            ("kt", "Period coeff. kt",       seism.get("kt",0.075), "—",  "Table 5-2",        "§5.1.2"),
            ("H",  "Building Height H",      seism.get("H",10.0),   "m",  "Total height",     "§5.1.2"),
            ("Ru", "Ductility Rμ",           seism.get("Ru",4.0),   "—",  "Table 5-2",        "§5.3"),
            ("Ou", "Overstrength Ωu",        seism.get("O_u",1.5),  "—",  "Table 5-2",        "§5.3"),
            ("Os", "Overstrength Ωs",        seism.get("O_s",1.5),  "—",  "Table 5-2",        "§5.3"),
            ("al", "α (spectral shape)",     seism.get("alpha",2.25),"—", "Soil Table 4-1",   "§4.1.3"),
            ("Tc", "Tc [corner period]",     seism.get("Tc",2.0),   "s",  "Soil Table 4-1",   "§4.1.3"),
            ("Td", "Td [long-period corner]",seism.get("Td",5.0),   "s",  "Soil Table 4-1",   "§4.1.3"),
            ("Ori","Orientation", "Parallel" if seism.get("is_parallel") else "Non-Parallel", "—", "User Input", "§3.6"),
            ("Sn", "Include Snow", "Yes" if seism.get("include_snow") else "No", "—", "User Input", "§3.6"),
        ]:
            ROW[key] = r
            _lbl(ws2,r,1,label,bold=True,bg=C_ALT_ROW if len(ROW)%2==0 else None)
            _inp(ws2,r,2,val, fmt="0.0000")
            _lbl(ws2,r,3,unit); _lbl(ws2,r,4,note); _lbl(ws2,r,5,clause)
            r+=1
        r+=1

        # ── SECTION B: Period Calculation (LIVE FORMULAS) ──────────────────
        _sec(ws2, r, "B. FUNDAMENTAL PERIOD  T = 1.25 × kt × H^(3/4)  (NBC 105:2025 §5.1.2–3)", end_col=5); r+=1
        _hdr(ws2,r,1,"Calculation"); _hdr(ws2,r,2,"Formula Cell"); _hdr(ws2,r,3,"Result")
        _hdr(ws2,r,4,"Formula Shown"); _hdr(ws2,r,5,"Clause"); r+=1

        ROW["T_approx"] = r
        _lbl(ws2,r,1,"T_approx = kt × H^(3/4)", bold=True)
        _fml(ws2,r,2,f"=B{ROW['kt']}*B{ROW['H']}^(3/4)", fmt="0.0000")
        _lbl(ws2,r,3,"s (approx)")
        _lbl(ws2,r,4,"kt × H^(3/4)")
        _lbl(ws2,r,5,"§5.1.2"); r+=1

        ROW["T"] = r
        _lbl(ws2,r,1,"T_design = 1.25 × T_approx", bold=True)
        _fml(ws2,r,2,f"=1.25*B{ROW['T_approx']}", fmt="0.0000")
        _lbl(ws2,r,3,"s (design)")
        _lbl(ws2,r,4,"1.25 × T_approx")
        _lbl(ws2,r,5,"§5.1.3  (amplification ×1.25)"); r+=1
        r+=1

        # ── SECTION C: Spectral Shape Ch(T) (LIVE FORMULAS) ────────────────
        _sec(ws2, r, "C. SPECTRAL SHAPE FACTOR Ch(T)  (NBC 105:2025 §4.1.2)", end_col=5); r+=1
        _hdr(ws2,r,1,"Parameter"); _hdr(ws2,r,2,"Formula Cell"); _hdr(ws2,r,3,"Unit")
        _hdr(ws2,r,4,"Formula (3-zone: flat → velocity → displacement)"); _hdr(ws2,r,5,"Clause"); r+=1

        ROW["Ch"] = r
        _lbl(ws2,r,1,"Ch(T)", bold=True)
        # 3-zone IF formula:
        # flat:     T < Tc  → α
        # velocity: Tc ≤ T < Td → α × Tc/T
        # disp:     T ≥ Td → α × Tc × Td / T²
        ch_formula = (f"=IF(B{ROW['T']}<B{ROW['Tc']},"
                      f"B{ROW['al']},"
                      f"IF(B{ROW['T']}<B{ROW['Td']},"
                      f"B{ROW['al']}*B{ROW['Tc']}/B{ROW['T']},"
                      f"B{ROW['al']}*B{ROW['Tc']}*B{ROW['Td']}/B{ROW['T']}^2))")
        _fml(ws2,r,2,ch_formula, fmt="0.0000")
        _lbl(ws2,r,3,"—")
        _lbl(ws2,r,4,"IF(T<Tc,α, IF(T<Td, α×Tc/T, α×Tc×Td/T²))")
        _lbl(ws2,r,5,"§4.1.2 Eq.(4.1.2)"); r+=1

        ROW["CT"] = r
        _lbl(ws2,r,1,"C(T) = Ch(T) × Z × I", bold=True)
        _fml(ws2,r,2,f"=B{ROW['Ch']}*B{ROW['Z']}*B{ROW['I']}", fmt="0.0000")
        _lbl(ws2,r,3,"—"); _lbl(ws2,r,4,"Ch(T) × Z × I"); _lbl(ws2,r,5,"§4.1.1"); r+=1

        ROW["Cs"] = r
        _lbl(ws2,r,1,"Cs(T) = 0.20 × C(T)  [SLS]", bold=True)
        _fml(ws2,r,2,f"=0.20*B{ROW['CT']}", fmt="0.0000")
        _lbl(ws2,r,3,"—"); _lbl(ws2,r,4,"0.20 × C(T)"); _lbl(ws2,r,5,"§4.2"); r+=1

        ROW["Cv"] = r
        _lbl(ws2,r,1,"Cv(T) = 2/3 × Z  [Vertical]", bold=True)
        _fml(ws2,r,2,f"=(2/3)*B{ROW['Z']}", fmt="0.0000")
        _lbl(ws2,r,3,"—"); _lbl(ws2,r,4,"(2/3) × Z"); _lbl(ws2,r,5,"§4.3"); r+=1
        r+=1

        # ── SECTION D: Base Shear Coefficients (LIVE FORMULAS) ─────────────
        _sec(ws2, r, "D. BASE SHEAR COEFFICIENTS  (NBC 105:2025 §6.1)", end_col=5); r+=1
        _hdr(ws2,r,1,"Parameter"); _hdr(ws2,r,2,"Formula Cell"); _hdr(ws2,r,3,"Unit")
        _hdr(ws2,r,4,"Formula"); _hdr(ws2,r,5,"Clause"); r+=1

        ROW["Cd_ULS"] = r
        _lbl(ws2,r,1,"Cd(T) ULS = C(T) / (Rμ × Ωu)", bold=True)
        _fml(ws2,r,2,f"=B{ROW['CT']}/(B{ROW['Ru']}*B{ROW['Ou']})", fmt="0.0000")
        _lbl(ws2,r,3,"—"); _lbl(ws2,r,4,"C(T) / (Rμ × Ωu)"); _lbl(ws2,r,5,"§6.1.1"); r+=1

        ROW["Cd_SLS"] = r
        _lbl(ws2,r,1,"Cd(T) SLS = Cs(T) / Ωs", bold=True)
        _fml(ws2,r,2,f"=B{ROW['Cs']}/B{ROW['Os']}", fmt="0.0000")
        _lbl(ws2,r,3,"—"); _lbl(ws2,r,4,"Cs(T) / Ωs"); _lbl(ws2,r,5,"§6.1.2"); r+=1
        r+=1

        # ── SECTION E: Story Force Distribution ────────────────────────────
        sf = seism.get("story_forces", [])
        if sf:
            _sec(ws2, r, "E. STORY FORCE DISTRIBUTION  (NBC 105:2025 §6.3)", end_col=6); r+=1
            ws2.cell(row=r, column=1,
                value="Formula: Fi = V × Wi×hi^k / Σ(Wj×hj^k)").font = Font(
                name="Arial", size=9, italic=True, color="595959")
            ws2.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); r+=1
            for h, lbl in [(1,"Floor"),(2,"Wi (kN)"),(3,"hi (m)"),(4,"Wi×hi^k"),(5,"Fi (kN)"),(6,"Story Shear Vx")]:
                _hdr(ws2,r,h,lbl)
            r += 1
            sf_start = r
            for i, frow in enumerate(sf):
                _inp(ws2,r,1,frow["floor"], fmt="0", align="center")
                _inp(ws2,r,2,frow["W_kN"], fmt="#,##0.0")
                _inp(ws2,r,3,frow["h_m"],  fmt="0.00")
                _fml(ws2,r,4,f"=B{r}*C{r}^{seism.get('k',1.0):.4f}", fmt="#,##0.0")
                r+=1
            # Fi formulas (need Σ Wh^k first)
            sf_end = r - 1
            sum_wh_ref = f"SUM(D{sf_start}:D{sf_end})"
            # V_base estimate
            V_base = seism.get("V_base_kN", seism.get("Cd_ULS",0)*seism.get("W_seismic_kN",0))
            _inp(ws2, r, 1, "V_base [kN]", align="left")
            _inp(ws2, r, 2, round(V_base,2), fmt="#,##0.00")
            Vbase_cell = f"B{r}"; r+=1
            # Now add Fi and Vx in new columns (re-using rows sf_start to sf_end)
            ws2.column_dimensions["E"].width = 14
            ws2.column_dimensions["F"].width = 18
            for i_sf, row_sf in enumerate(range(sf_start, sf_end+1)):
                _fml(ws2, row_sf, 5,
                     f"={Vbase_cell}*D{row_sf}/{sum_wh_ref}", fmt="#,##0.00")
                _fml(ws2, row_sf, 6,
                     f"=SUM(E{row_sf}:E{sf_end})", fmt="#,##0.00")

            _lbl(ws2, r, 1, "V check (ΣFi):", bold=True)
            _fml(ws2, r, 5, f"=SUM(E{sf_start}:E{sf_end})", fmt="#,##0.00")
            r+=2

        # ── SECTION F: Load Combinations ───────────────────────────────────
        combos = seism.get("load_combos", [])
        if combos:
            _sec(ws2, r, "F. LOAD COMBINATIONS  (NBC 105:2025 §3.6 LSM)", end_col=5); r+=1
            _hdr(ws2,r,1,"Combination"); _hdr(ws2,r,2,"Formula")
            _hdr(ws2,r,3,"DL"); _hdr(ws2,r,4,"LL (λ)"); _hdr(ws2,r,5,"E/W"); r+=1
            for i, c in enumerate(combos):
                bg = C_ALT_ROW if i%2==0 else None
                if "EX_ULS_fac" in c and (c.get("EX_ULS_fac",0)!=0 or c.get("EY_ULS_fac",0)!=0):
                    e_str = "".join([
                        f"{c['EX_ULS_fac']:+.1f}×EX" if c.get("EX_ULS_fac",0)!=0 else "",
                        f"{c['EY_ULS_fac']:+.1f}×EY" if c.get("EY_ULS_fac",0)!=0 else ""
                    ])
                else:
                    e_str = f"{c['E_SLS_fac']:+.1f}×E_SLS" if c.get('E_SLS_fac',0)!=0 else "—"
                ll_s = f"λ={seism.get('lambda_ll',0.30):.2f}" if str(c['LL_fac'])=="λ" else f"{c['LL_fac']:.2f}"
                _lbl(ws2,r,1,c["label"],bold=True,bg=bg)
                _lbl(ws2,r,2,c["formula"],bg=bg)
                _lbl(ws2,r,3,str(c["DL_fac"]),bg=bg)
                _lbl(ws2,r,4,ll_s,bg=bg)
                _lbl(ws2,r,5,e_str,bg=bg)
                r+=1

        ws2.freeze_panes = "A6"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — BEAM with LIVE FORMULAS
    # ══════════════════════════════════════════════════════════════════════════
    if beam and mode == "detailed":
        ws3 = wb.create_sheet("Beam (NBC+IS)")

        for col_l, w in [("A",38),("B",18),("C",14),("D",32),("E",22)]:
            ws3.column_dimensions[col_l].width = w

        r = _proj_header(ws3, "Beam Design — NBC 105:2025 Priority + IS 456 Fallback (Live Formulas)")

        ws3.merge_cells(f"A{r}:E{r}")
        lc = ws3.cell(row=r, column=1,
            value="BLUE = inputs (editable)   |   GREEN = Excel formulas (auto-recalculate)")
        lc.font = Font(name="Arial", size=9, italic=True, color="595959"); r+=1

        _sec(ws3, r, "A. SECTION INPUTS  (edit blue cells)", end_col=5); r+=1
        _hdr(ws3,r,1,"Parameter"); _hdr(ws3,r,2,"Value"); _hdr(ws3,r,3,"Unit")
        _hdr(ws3,r,4,"Description"); _hdr(ws3,r,5,"Clause"); r+=1

        BR = {}
        for key, lbl, val, unit, desc, clause in [
            ("b",   "Width b",           beam.get("b",300),   "mm", "Beam width",            "§6.2"),
            ("D",   "Overall Depth D",   beam.get("D",500),   "mm", "Overall depth",          "§6.2"),
            ("cov", "Cover",             beam.get("cover",25),"mm", "Clear cover to stirrup", "§26.4"),
            ("dia", "Main bar Ø",        beam.get("main_dia",16),"mm","Main rebar diameter",  "—"),
            ("fck", "fck",               beam.get("fck",25),  "MPa","Concrete cube strength", "§6.1"),
            ("fy",  "fy",                beam.get("fy",415),  "MPa","Steel yield strength",   "§5.6"),
            ("Mu",  "Design Mu",         beam.get("Mu_design_kNm",0),"kN·m","Factored bending moment","§22.2"),
            ("Vu",  "Design Vu",         beam.get("Vu_design_kN",0), "kN", "Factored shear force",  "§22.5"),
        ]:
            BR[key] = r
            _lbl(ws3,r,1,lbl,bold=True)
            _inp(ws3,r,2,val, fmt="0.00")
            _lbl(ws3,r,3,unit); _lbl(ws3,r,4,desc); _lbl(ws3,r,5,clause)
            r+=1
        r+=1

        _sec(ws3, r, "B. DERIVED GEOMETRY  (live formulas)", end_col=5); r+=1
        _hdr(ws3,r,1,"Parameter"); _hdr(ws3,r,2,"Formula"); _hdr(ws3,r,3,"Unit")
        _hdr(ws3,r,4,"Formula Shown"); _hdr(ws3,r,5,"Clause"); r+=1

        BR["d"] = r
        _lbl(ws3,r,1,"Effective depth d", bold=True)
        _fml(ws3,r,2,f"=B{BR['D']}-B{BR['cov']}-B{BR['dia']}/2", fmt="0.00")
        _lbl(ws3,r,3,"mm"); _lbl(ws3,r,4,"D - cover - Ø/2"); _lbl(ws3,r,5,"§26.3"); r+=1

        BR["xu_max"] = r
        xu_lim = 0.48  # Fe415
        _lbl(ws3,r,1,"xu,max = 0.48d  (Fe415)", bold=True)
        _fml(ws3,r,2,f"=0.48*B{BR['d']}", fmt="0.00")
        _lbl(ws3,r,3,"mm"); _lbl(ws3,r,4,"0.48 × d"); _lbl(ws3,r,5,"Table F"); r+=1

        BR["Mulim"] = r
        _lbl(ws3,r,1,"Mu,lim", bold=True)
        _fml(ws3,r,2,
             f"=0.36*B{BR['fck']}*B{BR['b']}*B{BR['xu_max']}*(B{BR['d']}-0.42*B{BR['xu_max']})/1E6",
             fmt="0.000")
        _lbl(ws3,r,3,"kN·m")
        _lbl(ws3,r,4,"0.36·fck·b·xu,max·(d−0.42·xu,max)/1E6")
        _lbl(ws3,r,5,"§38.1"); r+=1

        BR["Ast_req"] = r
        _lbl(ws3,r,1,"Ast required  (Mu ≤ Mu,lim)", bold=True)
        # xu from Mu: 0.36fck·b·xu·(d−0.42xu) = Mu×1E6
        # Ast = 0.36·fck·b·xu / (0.87·fy)
        # For singly: approximate Ast ≈ Mu×1E6 / (0.87·fy·0.9·d)
        _fml(ws3,r,2,
             f"=MIN(B{BR['Mu']}*1E6/(0.87*B{BR['fy']}*0.9*B{BR['d']}), "
             f"0.36*B{BR['fck']}*B{BR['b']}*B{BR['xu_max']}/(0.87*B{BR['fy']}))",
             fmt="0.0")
        _lbl(ws3,r,3,"mm²")
        _lbl(ws3,r,4,"Mu×1E6/(0.87·fy·0.9·d)  [approx. for singly]")
        _lbl(ws3,r,5,"§38.1"); r+=1

        BR["Ast_min"] = r
        _lbl(ws3,r,1,"Ast minimum", bold=True)
        _fml(ws3,r,2,f"=0.85*B{BR['b']}*B{BR['d']}/B{BR['fy']}", fmt="0.0")
        _lbl(ws3,r,3,"mm²"); _lbl(ws3,r,4,"0.85·b·d/fy"); _lbl(ws3,r,5,"§26.5.1.1"); r+=1

        BR["Ld"] = r
        # τbd for M25 = 1.4 MPa
        _lbl(ws3,r,1,"Development length Ld", bold=True)
        tbd = {20:1.2, 25:1.4, 30:1.5, 35:1.7, 40:1.9}.get(int(beam.get("fck",25)), 1.4)
        _fml(ws3,r,2,f"=0.87*B{BR['fy']}*B{BR['dia']}/(4*{tbd})", fmt="0.0")
        _lbl(ws3,r,3,"mm")
        _lbl(ws3,r,4,f"0.87·fy·Ø/(4·τbd)  [τbd={tbd} MPa for M{beam.get('fck',25)}]")
        _lbl(ws3,r,5,"§26.2.1"); r+=1
        r+=1

        # Shear check (live)
        _sec(ws3, r, "C. SHEAR DESIGN  (IS 456:2000 §40)", end_col=5); r+=1
        _hdr(ws3,r,1,"Parameter"); _hdr(ws3,r,2,"Formula"); _hdr(ws3,r,3,"Unit")
        _hdr(ws3,r,4,"Note"); _hdr(ws3,r,5,"Clause"); r+=1

        BR["tau_v"] = r
        _lbl(ws3,r,1,"τv = Vu/(b×d)", bold=True)
        _fml(ws3,r,2,f"=B{BR['Vu']}*1000/(B{BR['b']}*B{BR['d']})", fmt="0.000")
        _lbl(ws3,r,3,"MPa"); _lbl(ws3,r,4,"Nominal shear stress"); _lbl(ws3,r,5,"§40.1"); r+=1

        # τc from engine (static — depends on pt)
        tau_c = beam.get("shear",{}).get("tau_c",0)
        _lbl(ws3,r,1,"τc (design shear strength)", bold=True)
        _inp(ws3,r,2,tau_c, fmt="0.000")
        _lbl(ws3,r,3,"MPa"); _lbl(ws3,r,4,f"Table 19, pt={100*beam.get('Ast_prov_mm2',1)/(beam.get('b',1)*beam.get('d_eff_mm',1)):.2f}%"); _lbl(ws3,r,5,"Table 19"); r+=1

        ws3.freeze_panes = "A6"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4 — FOUNDATION
    # ══════════════════════════════════════════════════════════════════════════
    if fndg and mode == "detailed":
        ws4 = wb.create_sheet("Foundation (IS 456)")
        for col_l, w in [("A",38),("B",20),("C",12),("D",30),("E",20)]:
            ws4.column_dimensions[col_l].width = w

        r = _proj_header(ws4, "Foundation Design — IS 456:2000 §34")
        ws4.merge_cells(f"A{r}:E{r}")
        lc = ws4.cell(row=r,column=1,
            value=f"Type: {fndg.get('footing_type', 'Isolated Footing')}   |   BLUE = inputs   |   GREEN = formulas")
        lc.font = Font(name="Arial",size=9,italic=True,color="595959"); r+=1

        _sec(ws4, r, "A. FOOTING TYPE & MATERIAL PROPERTIES", end_col=5); r+=1
        _hdr(ws4,r,1,"Parameter"); _hdr(ws4,r,2,"Value"); _hdr(ws4,r,3,"Unit")
        _hdr(ws4,r,4,"Note"); _hdr(ws4,r,5,"Clause"); r+=1

        FR = {}
        type_id = fndg.get("footing_type_id", 1)
        
        # Common properties for all types
        for key, lbl, val, unit, note, clause in [
            ("fck",  "Concrete grade fck",  fndg.get("fck", 25),          "MPa", "Design strength", "§6.1"),
            ("fy",   "Steel grade fy",      fndg.get("fy", 415),          "MPa", "Reinforcement",   "§5.6"),
            ("cov",  "Clear cover",         fndg.get("cover_mm", 50),     "mm",  "Durability",      "§26.4"),
            ("bar",  "Bar diameter",        fndg.get("bar_dia_mm", 12),   "mm",  "Main reinforcement", "—"),
        ]:
            FR[key] = r
            _lbl(ws4,r,1,lbl,bold=True)
            _inp(ws4,r,2,val,fmt="0.00")
            _lbl(ws4,r,3,unit); _lbl(ws4,r,4,note); _lbl(ws4,r,5,clause)
            r+=1
        r+=1

        # Type-specific inputs
        if type_id == 0:  # Concentric
            _sec(ws4, r, "B. CONCENTRIC FOOTING — INPUTS", end_col=5); r+=1
            _hdr(ws4,r,1,"Parameter"); _hdr(ws4,r,2,"Value"); _hdr(ws4,r,3,"Unit")
            _hdr(ws4,r,4,"Note"); _hdr(ws4,r,5,"Clause"); r+=1
            for key, lbl, val, unit, note, clause in [
                ("cc_b",   "Column width b",      fndg.get("cc_b_mm", 300)/1000,   "m", "Plan dimension", "—"),
                ("cc_D",   "Column depth D",      fndg.get("cc_D_mm", 400)/1000,   "m", "Plan dimension", "—"),
                ("cc_P",   "Service load P",      fndg.get("cc_P_kN", 600),        "kN", "Axial load", "—"),
                ("cc_sbc", "Safe bearing capacity",fndg.get("cc_sbc_kPa", 150),    "kN/m²", "Design SBC", "§3.8"),
            ]:
                FR[key] = r
                _lbl(ws4,r,1,lbl,bold=True)
                _inp(ws4,r,2,val,fmt="0.000")
                _lbl(ws4,r,3,unit); _lbl(ws4,r,4,note); _lbl(ws4,r,5,clause)
                r+=1
        elif type_id == 1:  # Eccentric
            _sec(ws4, r, "B. ECCENTRIC FOOTING — INPUTS", end_col=5); r+=1
            _hdr(ws4,r,1,"Parameter"); _hdr(ws4,r,2,"Value"); _hdr(ws4,r,3,"Unit")
            _hdr(ws4,r,4,"Note"); _hdr(ws4,r,5,"Clause"); r+=1
            for key, lbl, val, unit, note, clause in [
                ("ec_b",   "Column width b",      fndg.get("ec_b_mm", 300)/1000,   "m", "Plan dimension", "—"),
                ("ec_D",   "Column depth D",      fndg.get("ec_D_mm", 400)/1000,   "m", "Plan dimension", "—"),
                ("ec_P",   "Service load P",      fndg.get("ec_P_kN", 800),        "kN", "Axial load", "—"),
                ("ec_Mx",  "Moment Mx",          fndg.get("ec_Mx_kNm", 0),         "kN·m", "Bending X", "—"),
                ("ec_My",  "Moment My",          fndg.get("ec_My_kNm", 0),         "kN·m", "Bending Y", "—"),
                ("ec_sbc", "Safe bearing capacity",fndg.get("ec_sbc_kPa", 150),    "kN/m²", "Design SBC", "§3.8"),
            ]:
                FR[key] = r
                _lbl(ws4,r,1,lbl,bold=True)
                _inp(ws4,r,2,val,fmt="0.000")
                _lbl(ws4,r,3,unit); _lbl(ws4,r,4,note); _lbl(ws4,r,5,clause)
                r+=1
        else:  # Combined
            _sec(ws4, r, "B. COMBINED FOOTING — INPUTS", end_col=5); r+=1
            _hdr(ws4,r,1,"Parameter"); _hdr(ws4,r,2,"Value"); _hdr(ws4,r,3,"Unit")
            _hdr(ws4,r,4,"Note"); _hdr(ws4,r,5,"Clause"); r+=1
            for key, lbl, val, unit, note, clause in [
                ("cb1_b",  "Column 1 width b",    fndg.get("cb1_b_mm", 300)/1000,  "m", "Left column", "—"),
                ("cb1_D",  "Column 1 depth D",    fndg.get("cb1_D_mm", 400)/1000,  "m", "Left column", "—"),
                ("cb_P1",  "Column 1 load P1",    fndg.get("cb_P1_kN", 600),       "kN", "Left column", "—"),
                ("cb2_b",  "Column 2 width b",    fndg.get("cb2_b_mm", 300)/1000,  "m", "Right column", "—"),
                ("cb2_D",  "Column 2 depth D",    fndg.get("cb2_D_mm", 400)/1000,  "m", "Right column", "—"),
                ("cb_P2",  "Column 2 load P2",    fndg.get("cb_P2_kN", 700),       "kN", "Right column", "—"),
                ("cb_sp",  "Center-to-center spacing", fndg.get("cb_sp_m", 5.0),   "m", "Column spacing", "—"),
                ("cb_sbc", "Safe bearing capacity",fndg.get("cb_sbc_kPa", 120),    "kN/m²", "Design SBC", "§3.8"),
            ]:
                FR[key] = r
                _lbl(ws4,r,1,lbl,bold=True)
                _inp(ws4,r,2,val,fmt="0.000")
                _lbl(ws4,r,3,unit); _lbl(ws4,r,4,note); _lbl(ws4,r,5,clause)
                r+=1
        r+=1
        
        _sec(ws4, r, "C. DESIGN RESULTS", end_col=5); r+=1
        _hdr(ws4,r,1,"Parameter"); _hdr(ws4,r,2,"Value"); _hdr(ws4,r,3,"Unit")
        _hdr(ws4,r,4,"Note"); _hdr(ws4,r,5,"Clause"); r+=1
        
        for key, lbl, val, unit, note, clause in [
            ("L",    "Footing Length L",    fndg.get("L_mm",1800)/1000,   "m",   "Plan dimension", "§34.1"),
            ("B",    "Footing Breadth B",   fndg.get("B_mm",1800)/1000,   "m",   "Plan dimension", "§34.1"),
            ("D",    "Overall Depth D",     fndg.get("D_mm",450)/1000,    "m",   "Min 0.30m",      "§34.1.3"),
        ]:
            FR[key] = r
            _lbl(ws4,r,1,lbl,bold=True)
            _inp(ws4,r,2,val,fmt="0.000")
            _lbl(ws4,r,3,unit); _lbl(ws4,r,4,note); _lbl(ws4,r,5,clause)
            r+=1
        r+=1
        
        # Set up FR keys based on footing type for formulas section
        if type_id == 0:  # Concentric
            FR["P"] = FR["cc_P"]
            FR["SBC"] = FR["cc_sbc"]
            FR["cb"] = FR["cc_b"]
            FR["cD"] = FR["cc_D"]
        elif type_id == 1:  # Eccentric
            FR["P"] = FR["ec_P"]
            FR["SBC"] = FR["ec_sbc"]
            FR["cb"] = FR["ec_b"]
            FR["cD"] = FR["ec_D"]
        else:  # Combined
            FR["P"] = FR["cb_P1"]
            FR["SBC"] = FR["cb_sbc"]
            FR["cb"] = FR["cb1_b"]
            FR["cD"] = FR["cb1_D"]
        
        _sec(ws4, r, "B. DERIVED QUANTITIES  (live formulas)", end_col=5); r+=1
        _hdr(ws4,r,1,"Parameter"); _hdr(ws4,r,2,"Formula"); _hdr(ws4,r,3,"Unit")
        _hdr(ws4,r,4,"Formula Shown"); _hdr(ws4,r,5,"Clause"); r+=1

        FR["d"] = r
        _lbl(ws4,r,1,"Effective depth d", bold=True)
        _fml(ws4,r,2,f"=B{FR['D']}-0.05-0.012/2", fmt="0.000")
        _lbl(ws4,r,3,"m"); _lbl(ws4,r,4,"D − cover(50mm) − Ø/2(12mm)"); _lbl(ws4,r,5,"§34.1"); r+=1

        FR["A"] = r
        _lbl(ws4,r,1,"Footing area A", bold=True)
        _fml(ws4,r,2,f"=B{FR['L']}*B{FR['B']}", fmt="0.000")
        _lbl(ws4,r,3,"m²"); _lbl(ws4,r,4,"L × B"); _lbl(ws4,r,5,"§34.2"); r+=1

        FR["q"] = r
        _lbl(ws4,r,1,"Net soil pressure q", bold=True)
        _fml(ws4,r,2,f"=B{FR['P']}/B{FR['A']}", fmt="0.00")
        _lbl(ws4,r,3,"kN/m²"); _lbl(ws4,r,4,"P/A"); _lbl(ws4,r,5,"§34.2"); r+=1

        FR["qu"] = r
        _lbl(ws4,r,1,"Factored pressure qu", bold=True)
        _fml(ws4,r,2,f"=1.5*B{FR['q']}", fmt="0.00")
        _lbl(ws4,r,3,"kN/m²"); _lbl(ws4,r,4,"1.5 × q"); _lbl(ws4,r,5,"§5.3"); r+=1

        FR["oh_L"] = r
        _lbl(ws4,r,1,"Overhang in L direction", bold=True)
        _fml(ws4,r,2,f"=(B{FR['L']}-B{FR['cD']})/2", fmt="0.000")
        _lbl(ws4,r,3,"m"); _lbl(ws4,r,4,"(L−col_D)/2"); _lbl(ws4,r,5,"§34.4.1"); r+=1

        FR["MuL"] = r
        _lbl(ws4,r,1,"Moment Mu_L at col face", bold=True)
        _fml(ws4,r,2,f"=B{FR['qu']}*B{FR['B']}*B{FR['oh_L']}^2/2", fmt="0.00")
        _lbl(ws4,r,3,"kN·m"); _lbl(ws4,r,4,"qu × B × (oh)²/2"); _lbl(ws4,r,5,"§34.4.1"); r+=1

        FR["tau_v"] = r
        _lbl(ws4,r,1,"One-way τv", bold=True)
        crit_L = max(0.0, (float(fndg.get("L_mm",1800))-float(fndg.get("cD_mm",400)))/2000 - fndg.get("d_mm",400)/1000)
        _fml(ws4,r,2,
             f"=B{FR['qu']}*B{FR['B']}*MAX(0,(B{FR['L']}-B{FR['cD']})/2-B{FR['d']})*1000/(B{FR['B']}*1000*B{FR['d']}*1000)",
             fmt="0.000")
        _lbl(ws4,r,3,"MPa"); _lbl(ws4,r,4,"qu×B×crit_L/(B×d)"); _lbl(ws4,r,5,"§34.4.2a"); r+=1

        FR["tau_punch"] = r
        _lbl(ws4,r,1,"Punching τv", bold=True)
        _fml(ws4,r,2,
             f"=B{FR['qu']}*(B{FR['L']}*B{FR['B']}-(B{FR['cD']}+B{FR['d']})*(B{FR['cb']}+B{FR['d']}))*1000"
             f"/(2*(B{FR['cD']}+B{FR['d']}+B{FR['cb']}+B{FR['d']})*1000*B{FR['d']}*1000)",
             fmt="0.000")
        _lbl(ws4,r,3,"MPa"); _lbl(ws4,r,4,"qu×(A−Aex)/(b0×d)"); _lbl(ws4,r,5,"§31.6"); r+=1

        FR["tau_c_punch"] = r
        import math as _math
        fck_val = 25
        _lbl(ws4,r,1,"Punching τc = 0.25√fck", bold=True)
        _fml(ws4,r,2,f"=0.25*SQRT(B{FR['fck']})", fmt="0.000")
        _lbl(ws4,r,3,"MPa"); _lbl(ws4,r,4,"0.25 × √fck"); _lbl(ws4,r,5,"§31.6.3.1"); r+=1

        r+=1
        _sec(ws4, r, "C. SAFETY CHECKS  (live — update automatically when inputs change)", end_col=5); r+=1
        _hdr(ws4,r,1,"Check"); _hdr(ws4,r,2,"Computed"); _hdr(ws4,r,3,"Limit")
        _hdr(ws4,r,4,"Formula Check"); _hdr(ws4,r,5,"Result"); r+=1

        for ch_lbl, computed_fml, limit_val, limit_fml, ok_fml in [
            ("q ≤ SBC",
             f"=TEXT(B{FR['q']},\"0.00\")&\" kN/m²\"",
             fndg.get("SBC_used_kPa",150),
             "SBC [kN/m²]",
             f"=IF(B{FR['q']}<=B{FR['SBC']},\"OK  ✓\",\"REVISE  ✗\")"),
            ("One-way shear",
             f"=TEXT(B{FR['tau_v']},\"0.000\")&\" MPa\"",
             fndg.get("tau_c_L",0.36),
             f"τc={fndg.get('tau_c_L',0.36):.3f} MPa",
             f"=IF(B{FR['tau_v']}<={fndg.get('tau_c_L',0.36)},\"OK  ✓\",\"REVISE  ✗\")"),
            ("Punching shear",
             f"=TEXT(B{FR['tau_punch']},\"0.000\")&\" MPa\"",
             fndg.get("tau_c_punch",1.25),
             f"=TEXT(0.25*SQRT(B{FR['fck']}),\"0.000\")&\" MPa\"",
             f"=IF(B{FR['tau_punch']}<=0.25*SQRT(B{FR['fck']}),\"OK  ✓\",\"REVISE  ✗\")"),
            ("Min depth ≥ 300mm",
             f"=TEXT(B{FR['D']}*1000,\"0\")&\" mm\"",
             "300 mm",
             "IS 456 §34.1.3",
             f"=IF(B{FR['D']}>=0.3,\"OK  ✓\",\"REVISE  ✗\")"),
        ]:
            _lbl(ws4,r,1,ch_lbl,bold=True)
            ws4.cell(row=r,column=2,value=computed_fml).font = Font(name="Arial",size=10)
            ws4.cell(row=r,column=2).border = _thin()
            _lbl(ws4,r,3,str(limit_val))
            ws4.cell(row=r,column=4,value=limit_fml).font = Font(name="Arial",size=10,color="595959")
            ws4.cell(row=r,column=4).border = _thin()
            c5 = ws4.cell(row=r,column=5,value=ok_fml)
            c5.font = Font(name="Arial",size=10,bold=True,color=C_OK_FG)
            c5.border = _thin()
            r+=1

        ws4.freeze_panes = "A6"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 5 — NOTES & DISCLAIMER
    # ══════════════════════════════════════════════════════════════════════════
    ws_n = wb.create_sheet("Notes")
    ws_n.column_dimensions["A"].width = 90
    c = ws_n.cell(row=1, column=1, value="NOTES & DISCLAIMER")
    c.font = Font(name="Arial",size=13,bold=True,color=C_TITLE_FG)
    c.fill = PatternFill("solid", fgColor=C_TITLE_BG)
    ws_n.row_dimensions[1].height = 26
    for i, (lbl, note) in enumerate([
        ("Colour convention:", "BLUE = editable inputs  |  GREEN = live Excel formulas (recalculate on change)"),
        ("Seismic formulas:",  "NBC 105:2025 §4.1.2 Ch(T): 3-zone formula — flat plateau → velocity-sensitive → displacement-sensitive"),
        ("Period formula:",    "T = 1.25 × kt × H^(3/4) per NBC 105:2025 §5.1.2–5.1.3 (amplification factor 1.25)"),
        ("Base shear:",        "Cd_ULS = C(T)/(Rμ×Ωu)  |  Cd_SLS = Cs(T)/Ωs  (NBC 105:2025 §6.1)"),
        ("Seismic Wi in app:", "Story Wi = DL + λ·LL with λ=0.30 for stories and λ=0.00 at top story (app setting)."),

        ("Seismic SBC:",       "For seismic footing designs: SBC × 1.5 (NBC 105:2025 §3.8)"),
        ("Beam Ast approx:",   "Ast formula in Excel uses lever-arm approximation (0.9d). Exact iterative solution available in app."),
        ("Disclaimer:",        "For reference only. All results must be independently verified by a qualified structural engineer."),
        ("Software:",          f"{APP_NAME} {APP_VERSION}  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"),
    ], start=2):
        ws_n.cell(row=i, column=1,
            value=f"{lbl}  {note}").font = Font(name="Arial", size=10,
            italic=(lbl=="Disclaimer:"))

    wb.save(output_path)
    return output_path
